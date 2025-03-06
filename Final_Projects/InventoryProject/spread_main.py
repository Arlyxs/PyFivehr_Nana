import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["Sheet1"]

# :- number of products per supplier
# :- (using dictionary key value pairs suppleir:product)
products_per_supplier = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row):  # last row excluded
    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        # print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

print(products_per_supplier)

# calculation of total value  of inventory per supplier dict supplier:ttl_prod_val
ttl_val_per_supplier = {}

for product_row in range(2, product_list.max_row):

    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    print(inventory)
    print(price)
    print(product_row)

    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in ttl_val_per_supplier:
        current_product_value_supplier = ttl_val_per_supplier[supplier_name]
        ttl_val_per_supplier[supplier_name] = (
            current_product_value_supplier + product_list.cell(product_row, 5).value
        )
    else:
        ttl_val_per_supplier[supplier_name] = inventory * price
        print("no value founmd)")


print(ttl_val_per_supplier)
