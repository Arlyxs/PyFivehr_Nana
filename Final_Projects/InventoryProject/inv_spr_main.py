from multiprocessing import Value
import openpyxl

# :- inventory per supplier | dict supplier:ttl_prod_val
inv_file = openpyxl.load_workbook(
    "/home/lyxc/Documents/8_IT_Development/2_CodingLessons/BackEnd/Python/Projects/UTube/TechWorld/PyFivehr_Nana/Final_Projects/InventoryProject/inventory.xlsx",
    data_only=True,
)
product_list = inv_file["Sheet1"]
# :- calculates products per supplier
# :- calculation of total value  of inventory per supplier
# :- products with inventory less than 10
products_per_supplier = {}
ttl_val_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row):  # last row excluded
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    supplier_row_val = inventory * price
    # supplier_row_val = product_list.cell(product_row, 5).value
    product_num = product_list.cell(product_row, 1).value

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a new: prod_supplier")
        products_per_supplier[supplier_name] = 1

    if supplier_name in ttl_val_per_supplier:
        current_product_value_supplier = ttl_val_per_supplier[supplier_name]
        ttl_val_per_supplier[supplier_name] = (
            current_product_value_supplier + supplier_row_val
        )
    else:
        ttl_val_per_supplier[supplier_name] = supplier_row_val
        print(f"adding new ttl_val_supplier: {supplier_name} ")

    if inventory < 10:
        products_under_10_inv[product_num] = inventory

# :- eureka print instructions
print(products_per_supplier)
print("Working: dictionary written_(npi)")
print(ttl_val_per_supplier)
print("Working: with cell multiplication")
print(
    f"The products with inventory less than ten are\n Product number: Quantity {products_under_10_inv}"
)
# ****************************************************************
