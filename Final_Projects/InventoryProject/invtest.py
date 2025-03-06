import openpyxl


inv_file = openpyxl.load_workbook(
    "/home/lyxc/Documents/8_IT_Development/2_CodingLessons/BackEnd/Python/Projects/UTube/TechWorld/PyFivehr_Nana/Final_Projects/InventoryProject/inventory.xlsx",
    data_only=True,
)


product_list = inv_file["Sheet1"]
"""
# :- number of products per supplier
# :- (using dictionary key value pairs suppleir:product)
products_per_supplier = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row):  # last row excluded
    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        # print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

print(products_per_supplier) """

# calculation of total value  of inventory per supplier dict supplier:ttl_prod_val
ttl_val_per_supplier = {}

for product_row in range(2, product_list.max_row):  # from first to last(incl)

    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    supplier_name = product_list.cell(product_row, 4).value

    supp_row_val = product_list.cell(product_row, 5).value

    print(product_list.max_row)

    if supplier_name in ttl_val_per_supplier:
        current_product_value_supplier = ttl_val_per_supplier[supplier_name]
        ttl_val_per_supplier[supplier_name] = (
            current_product_value_supplier + supp_row_val
        )
    else:
        ttl_val_per_supplier[supplier_name] = supp_row_val
        print(f"adding a new supplier {supplier_name} ")

print(ttl_val_per_supplier)


import openpyxl


def multiply_cells(cell1, cell2):
    if cell1.value is not None and cell2.value is not None:
        if isinstance(cell1.value, (int, float)) and isinstance(
            cell2.value, (int, float)
        ):
            return cell1.value * cell2.value
        else:
            return "Non-numeric values"  # Or handle as needed
    else:
        return "One or more cells are empty"


workbook = openpyxl.load_workbook("your_excel_file.xlsx")
sheet = workbook.active
cell_a1 = sheet["A1"]
cell_b1 = sheet["B1"]

result = multiply_cells(cell_a1, cell_b1)
print(result)

workbook.close()
# *********************************************************************

import openpyxl


def multiply_cells_convert(cell1, cell2):
    if cell1.value is not None and cell2.value is not None:
        try:
            value1 = float(cell1.value)
            value2 = float(cell2.value)
            return value1 * value2
        except ValueError:
            return "Cannot convert to number"
    else:
        return "One or more cells are empty"


workbook = openpyxl.load_workbook("your_excel_file.xlsx")
sheet = workbook.active
cell_a1 = sheet["A1"]
cell_b1 = sheet["B1"]

result = multiply_cells_convert(cell_a1, cell_b1)
print(result)

workbook.close()

# **********************************************************************

import openpyxl

workbook = openpyxl.load_workbook("your_excel_file.xlsx")
sheet = workbook.active

cell_a1 = sheet["A1"]
cell_b1 = sheet["B1"]

if (
    cell_a1.value is not None
    and isinstance(cell_a1.value, (int, float))
    and cell_b1.value is not None
    and isinstance(cell_b1.value, (int, float))
):
    result = cell_a1.value * cell_b1.value
    print(result)
else:
    print("One or both cells are not numbers or are empty.")

workbook.close()
